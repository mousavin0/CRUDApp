from sqlite_commands import prepare_relational_database
from sqlite_commands import add_user
from sqlite_commands import is_username_unique
from sqlite_commands import user_exists
# to avoid revealing password in terminal
import getpass
#to encode password before saving in database
import bcrypt
#to check if files exist
import os
#to writerow in csv
import csv

from openpyxl import Workbook
from datetime import datetime


from mongodb_commands import create_post



login_log_file = 'user_login.csv'
login_history_file = 'login_history.xlsx'


def get_nonempty_input(fieldname, is_password=False):
    prompt = f'Ange ditt {fieldname}: '
    while True:
        if is_password:
            user_input = getpass.getpass(prompt)
        else:
            user_input = input(prompt)
        if user_input.strip():  # Check if the input is not empty after stripping whitespace
            return user_input.strip()
        else:
            print(f'{fieldname} får inte vara tom.')


def login_menu():
    while True:
        username = get_nonempty_input("användarnamn")
        password = get_nonempty_input("lösenord",is_password=True)
        if user_exists(username,password):
            with open(login_log_file,'a') as f:    
                writer = csv.writer(f)
                dt=datetime.now()
                writer.writerow([username,dt.year,dt.month,dt.day,dt.hour,dt.minute,dt.second])
            print(f'Inloggad som {username}: ')
            while True:
                usermenu = input("""Välj ett alternative:
                            1. Posta Meddelande 
                            2. Söka Meddelende 
                            3. Uppdatera mitt konto
                            4. Logga ut \n """)
                if usermenu == '1':
                    create_post_menu(username)
                elif usermenu == '2':
                    pass
                elif usermenu == '3':
                    pass
                elif usermenu == '4':
                    return
                else:
                    print('Ogiltig Val!')
        else:
            print("Autentisering misslyckades! Försök igen!")

def create_account_menu():
    firstname = get_nonempty_input("förnamn")
    lastname = get_nonempty_input("efternamn")
    while True:
        username = get_nonempty_input("användarnamn: ")
        if not is_username_unique(username):
            print("Användarnamnet är redan upptaget. Välj en annan.")
        else:
            break
    password = get_nonempty_input("lösenord",is_password=True)
    address = input("Ange ditt adress: ")
    telephone = input("Ange ditt telefonnummer: ")

    password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
    add_user(firstname,lastname,username,password_hash,address,telephone)


def create_post_menu(username):
    post_dic = {}
    post_dic['username'] = username
    post_dic['title'] = get_nonempty_input("rubrik")
    post_dic['message'] = get_nonempty_input("meddelande")
    image = input("Ange din bild: ")
    video = input("Ange din video: ")
    if image:
        post_dic['image'] = image
    if video:
        post_dic['video'] = video
    links = []
    while True:
        print('Skriv \'avsluta\' eller tryck bara på enter för att slutföra tillägget!')
        link = input("Ange din länk: ")
        if link == 'avsluta' or not link:
            #post
            if links:
                post_dic['links'] = links
            create_post(post_dic)
            break
        else:
            links.append(link)


def prepare_log_files():
    if not os.path.exists(login_log_file):
        header = ['username', 'year', 'month', 'day','hour','minute','second']
        with open(login_log_file, 'w') as f:
            writer = csv.writer(f)
            writer.writerow(header)
    if not os.path.exists(login_history_file):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'number_of_logins'
        ws['B1'] = 'year'
        ws['C1'] = 'month'
        ws['D1'] = 'day'
        ws['E1'] = 'hour'
        wb.save(login_history_file)



if __name__ == "__main__":
    prepare_relational_database()
    prepare_log_files()
    while True:
        menu = input("""Välj ett alternative:
                     1. Bli medlem 
                     2. Logga in 
                     3. Stäng Appen! \n """)

        if menu == '1':
            create_account_menu()
        elif menu == '2':
            login_menu()
        elif menu == '3':
            break
        else:
            print("Ingen gilgit inmatning! Försök igen!") 