

# to avoid revealing password in terminal
import getpass

#to check if files exist
import os
#to writerow in csv
import csv

from openpyxl import Workbook
from constants import login_log_file, login_summary_file

from datetime import datetime

from sqlite_commands import add_login, get_user_id, get_login_summary

from openpyxl import load_workbook

def get_nonempty_input(fieldname, is_password=False):
    prompt = f'Ange ditt {fieldname}: '
    while True:
        if is_password:
            user_input = getpass.getpass(prompt)
        else:
            user_input = input(prompt)
        if user_input.strip():  # Check if the input is not empty after stripping whitespace
            return user_input.strip()
        else:
            print(f'{fieldname} får inte vara tom.')





def append_to_log(username):
    prepare_log_file()
    with open(login_log_file,'a') as f:    
        writer = csv.writer(f)
        dt=datetime.now()
        user_id = get_user_id(username)
        writer.writerow([user_id,dt.year,dt.month,dt.day,dt.hour,dt.minute,dt.second])


def prepare_log_file():
    if not os.path.exists(login_log_file):
        header = ['user_id', 'year', 'month', 'day','hour','minute','second']
        with open(login_log_file, 'w') as f:
            writer = csv.writer(f)
            writer.writerow(header)


def prepare_summary_file():
        if not os.path.exists(login_summary_file):
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'number_of_logins'
            ws['B1'] = 'year'
            ws['C1'] = 'month'
            ws['D1'] = 'day'
            ws['E1'] = 'hour'
            wb.save(login_summary_file)
            wb.close()


# def append_to_summary(rows):
#         for row in rows:
#         if not os.path.exists(login_summary_file):
#             wb = Workbook()
#             ws = wb.active
#             ws['A1'] = 'number_of_logins'
#             ws['B1'] = 'year'
#             ws['C1'] = 'month'
#             ws['D1'] = 'day'
#             ws['E1'] = 'hour'
#             wb.save(login_summary_file)




def read_from_log_file_write_to_table():
    prepare_log_file()
    with open(login_log_file,'r') as f:
        reader_obj = csv.DictReader(f) 
        for row in reader_obj: 
            user_id = int(row['user_id'])
            add_login(user_id,int(row['year']),int(row['month']),int(row['day']),
                            int(row['hour']),int(row['minute']),int(row['second']))
            

def update_summary_file():
    prepare_summary_file()
    wb = load_workbook(login_summary_file)
    ws = wb.worksheets[0]
    for row in get_login_summary():
        ws.append(row)

    wb.save(login_summary_file)
    wb.close()

